#!/usr/bin/env python3
"""
excel_tables.py — CDC 1.0.0
Writes styled Excel workbooks (.xlsx) for all four analysis phases.
One workbook per phase; one sheet per table.

Author: Nadeem Khan, INRS-CAFSB
"""

from pathlib import Path
import pandas as pd
import numpy as np

# openpyxl_styling
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, Reference
import warnings
warnings.filterwarnings("ignore")

# ── Colour_palette
H_FILL  = "2C5F8A"
H_FONT  = "FFFFFF"
ALT_ODD  = "F7F9FC"
ALT_EVEN = "FFFFFF"
SIG_FILL = "FDECEA"
GOOD_FILL= "EAF8F0"
WARN_FILL= "FEF9E7"
BORDER_C = "DDDDDD"

def _thin_border():
    side = Side(style="thin", color=BORDER_C)
    return Border(left=side, right=side, top=side, bottom=side)

def _header_fill():
    return PatternFill("solid", fgColor=H_FILL)

def _alt_fill(row_idx: int):
    color = ALT_ODD if row_idx % 2 == 1 else ALT_EVEN
    return PatternFill("solid", fgColor=color)

def _fmt_val(v):
    """Return display value for a cell."""
    if isinstance(v, float):
        if np.isnan(v): return "—"
        if 0 < abs(v) < 0.001: return f"{v:.3e}"
        return round(v, 5)
    return v


def _write_sheet(ws, df: pd.DataFrame, sheet_title: str,
                 sig_col: str = None, sig_thresh: float = 0.05,
                 highlight_cols: dict = None):
    """
    Write a styled DataFrame to an openpyxl worksheet.
    sig_col       : column name whose values trigger row highlight when < sig_thresh
    highlight_cols: dict {col_name: ("red"|"green"|"yellow")} for cell-level coloring
    """
    if df is None or df.empty:
        ws.append(["No data available."])
        return

    df = df.copy()
    cols = df.columns.tolist()
    n_cols = len(cols)

    # ── Header_row ───────────────────────────────────────────────────────────
    header_font  = Font(bold=True, color=H_FONT, size=10, name="Calibri")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for j, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=j, value=col)
        cell.fill   = _header_fill()
        cell.font   = font = header_font
        cell.alignment = header_align
        cell.border = _thin_border()

    # ── Data_rows ────────────────────────────────────────────────────────────
    body_font  = Font(size=9, name="Calibri")
    mono_font  = Font(size=9, name="Courier New")
    num_align  = Alignment(horizontal="right")
    str_align  = Alignment(horizontal="left")

    # Numeric_columns
    num_cols = {c for c in cols if df[c].dtype in (float, int, "float64", "int64")}

    for i, row in enumerate(df.itertuples(index=False), start=2):
        ws.row_dimensions[i].height = 15
        row_vals = list(row)
        # Determine if this row should be highlighted (significant)
        sig_highlight = False
        if sig_col and sig_col in cols:
            v = row_vals[cols.index(sig_col)]
            if isinstance(v, float) and not np.isnan(v) and v < sig_thresh:
                sig_highlight = True

        for j, (col, val) in enumerate(zip(cols, row_vals), start=1):
            display = _fmt_val(val)
            cell    = ws.cell(row=i, column=j, value=display)
            cell.border = _thin_border()

            # Font + alignment
            if col in num_cols:
                cell.font      = mono_font
                cell.alignment = num_align
            else:
                cell.font      = body_font
                cell.alignment = str_align

            # Row_background
            if sig_highlight:
                cell.fill = PatternFill("solid", fgColor=SIG_FILL)
            else:
                cell.fill = _alt_fill(i - 1)

            # Cell-level_highlights
            if highlight_cols and col in highlight_cols:
                rule = highlight_cols[col]
                raw  = val if not isinstance(val, str) else np.nan
                if not (isinstance(raw, float) and np.isnan(raw)):
                    if rule == "red"   and isinstance(raw, float) and raw > 5:
                        cell.fill = PatternFill("solid", fgColor=SIG_FILL)
                    elif rule == "green" and isinstance(raw, float) and raw < 5:
                        cell.fill = PatternFill("solid", fgColor=GOOD_FILL)
                    elif rule == "yellow":
                        cell.fill = PatternFill("solid", fgColor=WARN_FILL)

    # ── Column_widths ─────────────────────────────────────
    for j, col in enumerate(cols, start=1):
        max_len = max(len(str(col)),
                      df[col].astype(str).str.len().max() if not df.empty else 0)
        ws.column_dimensions[get_column_letter(j)].width = min(max_len + 3, 28)

    # ── Freeze_top_row ────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Sheet_tab_colour ──────────────────────────────────────────────────────
    ws.sheet_properties.tabColor = H_FILL


def write_phase1_excel(corr_df: pd.DataFrame, hotelling_df: pd.DataFrame,
                        out_dir: str):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Cross-Trait Correlations"
    _write_sheet(ws1, corr_df, "Cross-trait correlations",
                 sig_col="pearson_p", sig_thresh=0.05)

    ws2 = wb.create_sheet("Hotelling T2")
    _write_sheet(ws2, hotelling_df, "Hotelling T2",
                 sig_col="p_value", sig_thresh=0.05)

    path = f"{out_dir}/phase1_hypothesis/Phase1_HypothesisTesting.xlsx"
    _ensure_dir(path)
    wb.save(path)
    print(f"    ↳ excel: {path}")


def write_phase2_excel(reg_df: pd.DataFrame, vif_df: pd.DataFrame,
                        out_dir: str):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Regression R2"
    _write_sheet(ws1, reg_df, "Regression comparison",
                 sig_col="OLS_Fp", sig_thresh=0.05)

    ws2 = wb.create_sheet("VIF Diagnostics")
    _write_sheet(ws2, vif_df, "VIF",
                 highlight_cols={"VIF": "red"})

    # Add_color-scale_conditional_formatting_to_R2_columns
    for col_name in ["OLS_R2","Ridge_R2","Lasso_R2","EN_R2","PCR_R2","PLS_R2"]:
        if reg_df is None or reg_df.empty or col_name not in reg_df.columns:
            continue
        col_idx = reg_df.columns.tolist().index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        n_rows = len(reg_df)
        ws1.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{n_rows+1}",
            ColorScaleRule(
                start_type="num", start_value=0,   start_color="FFFFFF",
                mid_type="num",   mid_value=0.5,   mid_color="AED6F1",
                end_type="num",   end_value=1.0,   end_color=H_FILL,
            )
        )

    path = f"{out_dir}/phase2_regression/Phase2_Regression.xlsx"
    _ensure_dir(path)
    wb.save(path)
    print(f"    ↳ excel: {path}")


def write_phase3_excel(mtc_df: pd.DataFrame, lambda_df: pd.DataFrame,
                        out_dir: str):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "MTC Summary"
    _write_sheet(ws1, mtc_df, "Multiple testing correction")

    ws2 = wb.create_sheet("Genomic Inflation")
    _write_sheet(ws2, lambda_df, "Lambda GC")

    # DataBar_on_lambda_gc
    if lambda_df is not None and not lambda_df.empty and "lambda_gc" in lambda_df.columns:
        col_idx = lambda_df.columns.tolist().index("lambda_gc") + 1
        cl      = get_column_letter(col_idx)
        n       = len(lambda_df)
        ws2.conditional_formatting.add(
            f"{cl}2:{cl}{n+1}",
            DataBarRule(start_type="num", start_value=0.9,
                        end_type="num",   end_value=1.5,
                        color="2C5F8A")
        )

    path = f"{out_dir}/phase3_mtc/Phase3_MultiTesting.xlsx"
    _ensure_dir(path)
    wb.save(path)
    print(f"    ↳ excel: {path}")


def write_phase4_excel(causal_df: pd.DataFrame, loo_df: pd.DataFrame,
                        out_dir: str):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "IV Causal Estimates"
    _write_sheet(ws1, causal_df, "IVW causal",
                 sig_col="p_IVW", sig_thresh=0.05)

    ws2 = wb.create_sheet("LOO Sensitivity")
    _write_sheet(ws2, loo_df, "Leave-one-out")

    # Color_scale_on_beta_IVW
    if causal_df is not None and not causal_df.empty and "beta_IVW" in causal_df.columns:
        col_idx = causal_df.columns.tolist().index("beta_IVW") + 1
        cl      = get_column_letter(col_idx)
        n       = len(causal_df)
        ws1.conditional_formatting.add(
            f"{cl}2:{cl}{n+1}",
            ColorScaleRule(
                start_type="num", start_value=-1, start_color="2C5F8A",
                mid_type="num",   mid_value=0,    mid_color="FFFFFF",
                end_type="num",   end_value=1,    end_color="E05A3A",
            )
        )

    path = f"{out_dir}/phase4_causal/Phase4_CausalInference.xlsx"
    _ensure_dir(path)
    wb.save(path)
    print(f"    ↳ excel: {path}")


def _ensure_dir(path: str):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
